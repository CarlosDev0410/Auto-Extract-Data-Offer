"""
Módulo para exportar dados para Excel com formatação profissional
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Exporta e formata planilhas Excel"""
    
    @staticmethod
    def export_data(df: pd.DataFrame, filename: str) -> bool:
        """
        Exporta DataFrame para Excel com formatação profissional
        
        Args:
            df: DataFrame com os dados
            filename: Nome do arquivo de saída
            
        Returns:
            bool: True se exportou com sucesso
        """
        try:
            print(f"📊 Gerando planilha {filename}...")
            
            # Exporta para Excel
            df.to_excel(filename, index=False, sheet_name='Oferta Relâmpago')
            
            # Aplica formatação
            ExcelExporter._format_excel(filename)
            
            print(f"✅ Planilha {filename} criada com sucesso!")
            return True
            
        except Exception as e:
            print(f"❌ Erro ao exportar para Excel: {e}")
            return False
    
    @staticmethod
    def _format_excel(filename: str):
        """
        Aplica formatação profissional ao arquivo Excel
        
        Args:
            filename: Nome do arquivo Excel
        """
        wb = load_workbook(filename)
        ws = wb.active
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Estilo das bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Aplica formatação no cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formatação das células de dados
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # Ajusta largura das colunas automaticamente
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo de 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congela a primeira linha (cabeçalho)
        ws.freeze_panes = 'A2'
        
        # Salva as alterações
        wb.save(filename)
